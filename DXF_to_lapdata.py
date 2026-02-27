import ezdxf
import matplotlib.pyplot as plt
import numpy as np
import os
import math
import turtle
import openpyxl
from openpyxl import Workbook
#=========================================(DXF檔案讀取)
def file_search(filename):#尋找檔案
    print("Opening file...")
    if not os.path.exists(filename):
        print("File not found:",filename)
    else:
        print("Success:",filename)

def entity_Calculate(msp):#計算線條
    line_count = 0
    arc_count = 0
    for entity in msp:
        if entity.dxftype() == "LINE":
            line_count += 1
        elif entity.dxftype() == "ARC":
            arc_count += 1

    print("Total LINEs:",line_count)
    print("Total ARCs: ",arc_count)    

def DXFfile_reader(msp):# 讀取檔案參數
    for entity in msp:
        if entity.dxftype() == "LINE":
            start = entity.dxf.start
            end = entity.dxf.end
            print("LINE:=================================")
            print(f"  Start point: ({start.x:.3f}, {start.y:.3f})")
            print(f"  End point:   ({end.x:.3f}, {end.y:.3f})")

        elif entity.dxftype() == "ARC":
            center = entity.dxf.center
            radius = entity.dxf.radius
            start_angle = entity.dxf.start_angle
            end_angle = entity.dxf.end_angle
            print("ARC:==================================")
            print(f"  Center:       ({center.x:.3f}, {center.y:.3f})")
            print(f"  Radius:       {radius:.3f}")
            print(f"  Start angle:  {start_angle:.3f}°")
            print(f"  End angle:    {end_angle:.3f}°")  

def draw_dxf(msp):# matplotlib 劃出路徑
    fig, ax = plt.subplots()
    
    for entity in msp:
        if entity.dxftype() == "LINE": #起始點座標與結束點座計算線段向量
            start = entity.dxf.start
            end = entity.dxf.end
            x_values = [start.x, end.x]
            y_values = [start.y, end.y]
            ax.plot(x_values, y_values, 'b')  # 藍線

        elif entity.dxftype() == "ARC":
            center = entity.dxf.center
            radius = entity.dxf.radius
            start_angle = math.radians(entity.dxf.start_angle)
            end_angle = math.radians(entity.dxf.end_angle)

            if end_angle < start_angle:# 處理角度跨越 360 的情況（例如 350>>10）
                end_angle += 2 * math.pi

            angles = np.linspace(start_angle, end_angle, num=100)#用半徑與三角函數算出座標並加入起始點偏移
            x_values = center.x + radius * np.cos(angles)
            y_values = center.y + radius * np.sin(angles)
            ax.plot(x_values, y_values, 'r')  # 紅弧線
            
    ax.set_aspect('equal')
    ax.set_title("DXF Track")
    plt.xlabel("X")
    plt.ylabel("Y")
    plt.grid(True)
    plt.show()
    
#========================================(資料排序)
class PathSegment:#線段元素儲存庫
    def __init__(self, type, start, end, center=None, radius=None, start_angle=None, end_angle=None,mid_angle = None,direction = None):
        self.type = type  # 'LINE' or 'ARC'
        self.start = start  # (x, y)
        self.end = end      # (x, y)
        self.center = center
        self.radius = radius
        self.start_angle = start_angle
        self.end_angle = end_angle
        self.mid_angle = mid_angle
        self.direction = direction

def tolerance(p1, p2, tol=1e-8): #容差判斷
    return (abs(p1[0] - p2[0])**2 + abs(p1[1] - p2[1])**2)**0.5 < tol #畢氏定理計算容差

def extract_segments(msp):#線段與弧資料儲存3
    # 這裡的邏輯和matplotlib 劃出路徑很像
    segments = []#保存清單

    for entity in msp:
        if entity.dxftype() == "LINE":
            start = entity.dxf.start
            end = entity.dxf.end
            seg = PathSegment(
                type='LINE',
                start=(start.x, start.y),
                end=(end.x, end.y)
            )
            segments.append(seg)

        elif entity.dxftype() == "ARC":
            center = entity.dxf.center
            radius = entity.dxf.radius
            start_angle_rad = math.radians(entity.dxf.start_angle)
            end_angle_rad = math.radians(entity.dxf.end_angle)

            # 轉成起點、終點座標
            start_x = center.x + radius * math.cos(start_angle_rad)
            start_y = center.y + radius * math.sin(start_angle_rad)
            end_x = center.x + radius * math.cos(end_angle_rad)
            end_y = center.y + radius * math.sin(end_angle_rad)
            if end_angle_rad < start_angle_rad:# 處理角度跨越 360 的情況（例如 350>>10）
                mid_angle_rad = (start_angle_rad+end_angle_rad+2 * math.pi)/2
            else:
                mid_angle_rad = (start_angle_rad+end_angle_rad)/2
            seg = PathSegment(
                type='ARC',
                start=(start_x, start_y),
                end=(end_x, end_y),
                center=(center.x, center.y),
                radius=radius,
                start_angle=start_angle_rad,
                end_angle=end_angle_rad,
                mid_angle=mid_angle_rad#確認方向用
            )
            segments.append(seg)

    return segments

def reverse_segment(seg):#反轉排序
    seg.start, seg.end = seg.end, seg.start
    if seg.type == 'ARC':
        seg.start_angle, seg.end_angle = seg.end_angle, seg.start_angle
    
def sort_segments(segments, tol=1e-8):#線段排序
    if not segments:
        return []

    sorted_list = [segments.pop(0)]  # 先選一個起點作為開頭

    while segments:
        last = sorted_list[-1]
        last_end = last.end #結束點 用來找下一段

        found = False
        for i, seg in enumerate(segments):# 比對點是否有在容差內有的話就加入清單
            if tolerance(seg.start, last_end, tol):
                sorted_list.append(segments.pop(i))
                found = True
                break
            elif tolerance(seg.end, last_end, tol):# 起始點和結束點是我自己定義的所以有可能反過來 所以可以倒轉方向比對看看
                reverse_segment(seg)
                sorted_list.append(segments.pop(i))
                found = True
                break

        if not found:
            print("The track is not continuous!!!")
            break

    return sorted_list

def print_segment_info(segments):#排序後資料參數
    print("\n====== Sorted Segment Information ======\n")
    for i, seg in enumerate(segments):
        print(f"Segment {i+1}: {seg.type}")
        print(f"  Start: {seg.start}")
        print(f"  End:   {seg.end}")
        if seg.type == 'ARC':
            print(f"  Center: {seg.center}")
            print(f"  Radius: {seg.radius}")
            print(f"  Angle:  {math.degrees(seg.start_angle):.1f}° -> {math.degrees(seg.end_angle):.1f}°")
            clockwise = seg.end_angle<seg.mid_angle and seg.mid_angle<seg.start_angle or seg.mid_angle>seg.end_angle and seg.mid_angle>seg.start_angle and seg.start_angle<seg.end_angle
            if clockwise:
                print(f"  clockwise")
            else:
                print(f"  Counterclockwise")
            #print(f"  mid Angle{math.degrees(seg.mid_angle):.1f}°")
        print()

def draw_segments(segments):# matplotlib 劃出路徑(主要是segments排序確認用)
    fig, ax = plt.subplots()

    for seg in segments:
        if seg.type == "LINE":
            x_values = [seg.start[0], seg.end[0]]
            y_values = [seg.start[1], seg.end[1]]
            ax.plot(x_values, y_values, 'b')  # 藍線

        elif seg.type == "ARC":
            center = seg.center
            radius = seg.radius
            start_angle = seg.start_angle
            end_angle = seg.end_angle
            clockwise = seg.end_angle<seg.mid_angle and seg.mid_angle<seg.start_angle or seg.mid_angle>seg.end_angle and seg.mid_angle>seg.start_angle and seg.start_angle<seg.end_angle
            if clockwise:
                if end_angle > start_angle:# 處理角度跨越 360 的情況（例如 350>>10）
                    start_angle += 2 * math.pi
                angles = np.linspace(end_angle, start_angle, num=100)
            else:
                if end_angle < start_angle:# 處理角度跨越 360 的情況（例如 350>>10）
                    end_angle += 2 * math.pi
                angles = np.linspace(start_angle, end_angle, num=100)

            x_values = center[0] + radius * np.cos(angles)
            y_values = center[1] + radius * np.sin(angles)
            ax.plot(x_values, y_values, 'r')  # 紅弧線
            
    for i, seg in enumerate(segments):
        x, y = seg.start
        ax.plot(x, y, 'go')  # 黑色圓點
        ax.text(x, y, f"{i+1}", fontsize=15, color='black', ha='center', va='center')
        
    ax.set_aspect('equal')
    ax.set_title("Track")
    plt.xlabel("X")
    plt.ylabel("Y")
    plt.grid(True)
    plt.show()

def rotate_segments(sorted_segments, start):#選擇起點重新排序
    #將已排序的 segments 從第 start 個開始，形成新的圓環順序
    n = len(sorted_segments)
    if not (0 <= start < n):
        raise ValueError(f"start 必須在 0 ~ {n-1} 之間")
    start = start-1
    return sorted_segments[start:] + sorted_segments[:start]

#==========================================(烏龜大小方向調整)
def size_scale(segments, scale=100):# 縮放圖面
    # 找範圍
    min_x = min(min(seg.start[0], seg.end[0]) for seg in segments)
    min_y = min(min(seg.start[1], seg.end[1]) for seg in segments)
    max_x = max(max(seg.start[0], seg.end[0]) for seg in segments)
    max_y = max(max(seg.start[1], seg.end[1]) for seg in segments)

    for seg in segments:
        # 縮放 start, end
        seg.start = ((seg.start[0] - min_x) * scale, (seg.start[1] - min_y) * scale)#將位置乘上scale係數
        seg.end   = ((seg.end[0] - min_x) * scale, (seg.end[1] - min_y) * scale)
        if seg.center != None:
            seg.center = ((seg.center[0] - min_x) * scale, (seg.center[1] - min_y) * scale)
            seg.radius = seg.radius*scale

def setup_turtle():#設定烏龜圖
    t = turtle.Turtle()
    t.speed(10)        # 1～10，或 "fastest"
    t.pensize(2)
    t.color("blue")
    t.penup()
    return t

def click_exit(x, y):#跳出烏龜(點一下就跳出來但我覺得其實按叉叉就好但是我這段先保留)
    print("Clicked, exiting...")
    turtle.bye()  # 關閉 turtle 視窗

def draw_with_turtle(segments):#畫出烏龜圖
    #turtle.onscreenclick(click_exit)  # 設定點擊事件不然每次都要跑完很浪費時間
    t = setup_turtle()
    try:
        
        for i, seg in enumerate(segments):
            print(i)
            if seg.type == "LINE":
                x, y = seg.start
                t.penup()
                t.goto(x, y)#去起始點
                t.pendown()
                t.goto(seg.end[0], seg.end[1])#開始畫

            elif seg.type == 'ARC':
                center = seg.center
                radius = seg.radius
                start_angle = math.degrees(seg.start_angle)
                end_angle = math.degrees(seg.end_angle)
                mid_angle = math.degrees(seg.mid_angle)

                # === 計算起點向量與切線方向 ===
                dx = seg.start[0] - center[0]
                dy = seg.start[1] - center[1]
                angle = math.degrees(math.atan2(dy, dx))

                # === 決定方向與 extent ===
                extent = (end_angle - start_angle) % 360
                
                clockwise = end_angle<mid_angle and mid_angle<start_angle or mid_angle>end_angle and mid_angle>start_angle and start_angle<end_angle
                # === 設定起點與方向 ===
                t.penup()
                t.goto(seg.start)
                if clockwise:
                    t.setheading(angle - 90)
                    t.pendown()
                    t.circle(-radius, 360 - extent)  # radius 為負，畫順時針
                    print("clockwise")
                else:
                    t.setheading(angle + 90)
                    t.pendown()
                    t.circle(radius, extent)
                    print("Counterclockwise")
        turtle.done()
    except turtle.Terminator:
        print("Is the direction correct?")

def Track_Route(segments,scale=100):#畫出正確比例烏龜圖
    size_scale(segments,scale)#縮放適當比例
    draw_with_turtle(segments)#跑出賽道圖形
    size_scale(segments,1/scale)#縮放適當比例

def reset_direction(segments):#重設方向
    segments.reverse()#重設順序
    for seg in segments:
        # 換起點與終點
        seg.start, seg.end = seg.end, seg.start

        # 換角度（僅針對 ARC）
        if seg.type == 'ARC':
            seg.start_angle, seg.end_angle = seg.end_angle, seg.start_angle

#==========================================(excel檔案輸出)
def excel_data(segments, filename ,workbook):
    if workbook is None:
        wb = openpyxl.Workbook()
    else:
        wb = workbook
    ws = wb.active
    ws.title = "Trackdata"

    headers = [
        "section", "Type", "Start_X", "Start_Y", "End_X", "End_Y",
        "Center_X", "Center_Y", "Radius", "Start_Angle (deg)", "End_Angle (deg)",
        "length","Direction"
    ]
    ws.append(headers)

    for i, seg in enumerate(segments):
        if seg.type == 'LINE':
            length = ((seg.start[0]-seg.end[0])**2+(seg.start[1]-seg.end[1])**2)**0.5#兩點算出直線距離
            row = [
                i + 1,
                "LINE",
                seg.start[0], seg.start[1],
                seg.end[0], seg.end[1],
                "", "", "", "", "",length,
                ""
            ]
        elif seg.type == 'ARC':
            # 計算方向：用中角判斷是否為順時針
            sa = seg.start_angle
            ea = seg.end_angle
            ma = seg.mid_angle
            clockwise = ea<ma and ma<sa or ma>ea and ma>sa and sa<ea
            direction = "Clockwise" if clockwise else "Counterclockwise"
            if clockwise:
                theta = abs((sa-ea)%(2 * math.pi))
                arc_length = seg.radius * (theta)#R*theta 計算弧長
            else:
                theta = abs((ea-sa)%(2 * math.pi))
                arc_length = seg.radius * (theta)#R*theta 計算弧長

            row = [
                i + 1,
                "ARC",
                seg.start[0], seg.start[1],
                seg.end[0], seg.end[1],
                seg.center[0], seg.center[1],
                seg.radius,
                math.degrees(sa),
                math.degrees(ea),
                arc_length,
                direction
            ]
        ws.append(row)

    wb.save(filename)
    print(f"main info exported to: {filename}")

def excel_export_summary(segments,filename, workbook,):
    ws = workbook.create_sheet(title="GeometrySummary")#同個檔案中的新表單

    headers = [
        "section", "Type", "Start_X", "Start_Y",
        "Center_X", "Center_Y", "Radius",
        "Length", "Direction"
    ]
    ws.append(headers)

    for i, seg in enumerate(segments):
        if seg.type == 'LINE':
            length = ((seg.start[0]-seg.end[0])**2+(seg.start[1]-seg.end[1])**2)**0.5
            row = [
                i + 1, "LINE",
                seg.start[0], seg.start[1],
                "", "", "", length, ""
            ]
        elif seg.type == 'ARC':
            # 計算弧長與方向
            sa = seg.start_angle
            ea = seg.end_angle
            ma = seg.mid_angle
            clockwise = ea<ma and ma<sa or ma>ea and ma>sa and sa<ea
            direction = "Clockwise" if clockwise else "Counterclockwise"
            if clockwise:
                theta = abs((sa-ea)%(2 * math.pi))
                arc_length = seg.radius * (theta)#R*theta 計算弧長
            else:
                theta = abs((ea-sa)%(2 * math.pi))
                arc_length = seg.radius * (theta)#R*theta 計算弧長

            row = [
                i + 1, "ARC",
                seg.start[0], seg.start[1],
                seg.center[0], seg.center[1],
                seg.radius,
                arc_length,
                direction
            ]
        ws.append(row)
    print(f"GeometrySummary info exported to: {filename}")

def export_all_to_excel(segments, filename):
    wb = Workbook()
    excel_data(segments, filename, wb)  # 傳入已存在的 wb
    excel_export_summary(segments,filename, wb)
    wb.save(filename)
    print(f"All data exported to {filename}")

#=====================================(執行)
def run_all_read_DXF(DXF_filename,excel_filename):#完整執行
    file_search(DXF_filename)#找檔案
    doc = ezdxf.readfile(DXF_filename)
    msp = doc.modelspace()
    DXFfile_reader(msp)#快速讀檔
    entity_Calculate(msp)#計算直線與彎道數量
    draw_dxf(msp)#預覽圖形

    segments = extract_segments(msp)
    print("total path:", len(segments))#計算總線段數量
    sorted_segments = sort_segments(segments.copy())  # 用 copy 避免改原資料並依照軌跡排序檔案
    print("sort segments:", len(sorted_segments))#線段排序
    draw_segments(sorted_segments)#二次確認資料是否有遺失，並且標記起始點
    start = int(input("start at : "))
    #start = 47
    setedstart_sorted_segments = rotate_segments(sorted_segments.copy(), start)#旋轉賽道到正確的起點
    print("Redefine the starting point")

    #Track_Route(setedstart_sorted_segments,100)#路線模擬
    print("checking direction")#
    direction = input("reset direction(Y/N):")
    if direction=="Y":
        print("reset_direction!!!")
        reset_direction(setedstart_sorted_segments)
    
    print("successfully sorted segments:", len(setedstart_sorted_segments))#線段排序成功
    draw_segments(setedstart_sorted_segments)#正確順序
    Track_Route(setedstart_sorted_segments,0.1)#路線模擬
    
    size_scale(setedstart_sorted_segments,0.1)#縮放真實賽道比例
    print_segment_info(setedstart_sorted_segments)#打印所有資料
    export_all_to_excel(setedstart_sorted_segments,excel_filename)#資料存檔
    return (setedstart_sorted_segments)


DXF_filename = "fsae_A_2025_track.dxf"
excel_filename = "fsae_A_2025_track.xlsx"
run_all_read_DXF(DXF_filename,excel_filename)
